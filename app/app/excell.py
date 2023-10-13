from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


columns=0
ref=""
batches={}
fixed_batches={}
peso_tela={};
combinations={}
batch_number=""

codigo=""
peso=""





def find_combination(search_value):
    for key,value in combinations.items():
       
       if set(value) == set(search_value):
           return True
     
    return False

##Paso 1 guardar columnas de la tabla 
def step_1(ws):
   for table in ws.tables.values():
       ref=table.ref
       
       columns= len(table.tableColumns)

##Paso 2 leer la fila de bath_number para guardar los IDs
def step_2(ws):
   for data in ws.iter_rows(max_col=columns, values_only=True,min_row=2):
       
       batch_number=str(data[0])
       
       if(batch_number not in batches.keys()):
           ##si el numero de lote (batch_number) no esta en la lista lo agregamos  utilzanod batch_numnber como 
           ##llave
           print("adding first key")
           ##data[1] es el tinte , agregamos una lista de tintes que represtan la combinacion
           batches[batch_number]=[data[1]]
           ##incluido guardamos peso tela en diccionario usando como llave el batch number
           peso_tela[batch_number]=data[2]
           
       else:
           ##si el batch ya existe solo agregamos el tinte a su combinacion
           batches[batch_number].append(data[1])
           print("key already found")
           
   
##Paso Creamos un nuevo dicionario llamado combinations 
def step_3():
    counter=0
    
    for key, value in batches.items():
        ##si la combinacion es unica la agregamos al diccionario
     if find_combination(value) !=True:
         
         if(counter<10):
             combinations[f"DT0{counter}"]=value
             
             
         else:  
             combinations[f"DT{counter}"]=value
         counter+=1
     else:
        ##No es una combinacion de tintes unicas
         print("se Repite")
 
 ##Creamos un diccionario  que contiene {batch_number,id_grupo}
def step_4():
    for key1, list_val in batches.items():
        for key,list_val2 in combinations.items():
            if set(list_val)==set(list_val2):
                fixed_batches[key1]=key;
    



def run_procedure(file,sheet,out):
      
   wb = load_workbook(filename=file,read_only=False)
   ws=wb[sheet]
   step_1(ws)
   step_2(ws)
   step_3()
   step_4()



  ##Creamos nuevo cuaderno
   workbook = openpyxl.Workbook()
   
   worksheet = workbook.active
   worksheet.title = "Lotes"
   
   #Encabezados
   header = ["lote","combinacion","peso tela"]
   worksheet.append(header)
   
   # Add data rows
   data = []
   for key,value in fixed_batches.items():
       
        ##Agergamos los teres encabezados de la hoa de otes
       data.append([key,value,peso_tela[key]])
   
   for row in data:
       worksheet.append(row)
   
   ##SEGUNDA HOJA
   worksheet2 = workbook.create_sheet("grupos")
   worksheet2.title = "Grupos"
   
   # Create a table header
   header2 = ["group id","colorante 1", "colorante 2", "coloreante 3", "colorante 4"]
   worksheet2.append(header2)
   
   # Add data rows
   data2 = []
   ##por cada combinacion unica
   for key,value in combinations.items():
        ## creamos sublista para almacenar cada tinte de la combinacion
       sub=[]
       sub.append(key),
       for item in value:
           sub.append(item)
       ##fin de algoritmo
    
       data2.append(sub)
       
   
   for row in data2:
       worksheet2.append(row)
   


   
   # Save the Excel file
   workbook.save(out)
   fixed_batches.clear()
   peso_tela.clear()
   combinations.clear()
   batches.clear()
   
   


   

